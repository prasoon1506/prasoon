import streamlit as st
import pandas as pd
import io
import openpyxl
import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.utils import get_column_letter

def get_visible_columns(ws):
    """
    Extract visible column names from an Excel worksheet
    
    Args:
        ws (openpyxl.worksheet.Worksheet): Excel worksheet
    
    Returns:
        list: List of visible column names
    """
    visible_columns = []
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        
        # Check column visibility
        try:
            # Some versions of openpyxl might handle column visibility differently
            is_hidden = ws.column_dimensions[column_letter].hidden if column_letter in ws.column_dimensions else False
        except:
            is_hidden = False
        
        # Get cell value safely
        try:
            cell_value = ws.cell(row=1, column=col).value
        except:
            cell_value = None
        
        # Add column if not hidden and has a value
        if not is_hidden and cell_value is not None:
            visible_columns.append(cell_value)
    
    return visible_columns

def read_visible_columns(file, sheet_name):
    """
    Read only visible columns from an Excel file
    
    Args:
        file (file-like object): Uploaded Excel file
        sheet_name (str): Name of the sheet to read
    
    Returns:
        pd.DataFrame: DataFrame with only visible columns
    """
    # Load workbook and worksheet
    wb = openpyxl.load_workbook(file, data_only=True, keep_vba=False)
    ws = wb[sheet_name]
    
    # Get visible column names
    visible_column_names = get_visible_columns(ws)
    
    # Read the DataFrame
    df = pd.read_excel(
        file, 
        sheet_name=sheet_name, 
        usecols=visible_column_names
    )
    
    return df

def preprocess_data(df):
    def safe_convert(value):
        """Safely convert value to string or return None"""
        try:
            # Convert to string, handling different types
            str_val = str(value).strip()
            return str_val if str_val else None
        except:
            return None

    # Convert to list and handle potential issues
    first_column_values = df.iloc[:, 0].tolist()
    
    # Preprocessing steps
    df = df.iloc[3:]
    df = df.reset_index(drop=True)
    
    # Remove rows containing 'Zone' (case-insensitive)
    df = df[~df.iloc[:, 0].astype(str).str.contains("Zone", case=False, na=False)]
    
    # Identify and remove rows around 'Region' rows
    region_indices = df[df.iloc[:, 0].astype(str).str.contains("Region", case=False, na=False)].index
    rows_to_remove = []
    for index in region_indices:
        rows_to_remove.extend([index - 2, index - 1, index])    
    df = df.drop(rows_to_remove)
    df = df.reset_index(drop=True)
    
    # Fill missing values in first column
    current_fill_value = None
    for i in range(len(df)):
        # Safely get and convert value
        value = safe_convert(df.iloc[i, 0])
        
        if value:
            current_fill_value = value
        elif current_fill_value is not None:
            df.iloc[i, 0] = current_fill_value
    
    # Remove 'All India' section
    all_india_mask = df.iloc[:, 0].astype(str).str.contains("All India", case=False, na=False)
    if all_india_mask.any():
        all_india_index = all_india_mask[all_india_mask].index[0]
        if all_india_index > 0:
            df = df.drop(all_india_index - 1)
            df = df.reset_index(drop=True)
    
    # Remove last two rows (typically summary or footer)
    df = df[:-2]
    
    return df

def process_and_merge(df, file_type):
    # Define column names based on file type
    if file_type == "Oct-Sep":
        column_names = ["Region Name", "Material Name", 
                        "Oct Trade Quantity", "Sep Trade Quantity",
                        "Oct Trade EBITDA", "Sep Trade EBITDA", "Increase in Trade EBITDA",
                        "Oct Non-Trade Quantity", "Sep Non-Trade Quantity", 
                        "Oct Non-Trade EBITDA", "Sep Non-Trade EBITDA", "Increase in Non-Trade EBITDA", 
                        "Oct Total Quantity", "Sep Total Quantity", 
                        "Oct Total EBITDA", "Sep Total EBITDA", "Increase in Total EBITDA"]
    elif file_type == "Sep-Aug":
        column_names = ["Region Name", "Material Name", 
                        "Sep Trade Quantity", "Aug Trade Quantity",
                        "Sep Trade EBITDA", "Aug Trade EBITDA", "Increase in Trade EBITDA",
                        "Sep Non-Trade Quantity", "Aug Non-Trade Quantity", 
                        "Sep Non-Trade EBITDA", "Aug Non-Trade EBITDA", "Increase in Non-Trade EBITDA", 
                        "Sep Total Quantity", "Aug Total Quantity", 
                        "Sep Total EBITDA", "Aug Total EBITDA", "Increase in Total EBITDA"]
    elif file_type == "Aug-Jul":
        column_names = ["Region Name", "Material Name", 
                        "Aug Trade Quantity", "Jul Trade Quantity",
                        "Aug Trade EBITDA", "Jul Trade EBITDA", "Increase in Trade EBITDA",
                        "Aug Non-Trade Quantity", "Jul Non-Trade Quantity", 
                        "Aug Non-Trade EBITDA", "Jul Non-Trade EBITDA", "Increase in Non-Trade EBITDA", 
                        "Aug Total Quantity", "Jul Total Quantity", 
                        "Aug Total EBITDA", "Jul Total EBITDA", "Increase in Total EBITDA"]
    
    # Select only the columns present in the DataFrame
    available_columns = [col for col in column_names if col in df.columns]
    df = df[available_columns]
    
    # Assign column names
    df.columns = available_columns

    # Create separate DataFrames for Total and Non-Total rows
    total_mask = df.iloc[:, 0].str.contains("Total", case=False, na=False)
    df_total = df[total_mask]
    df_no_total = df[~total_mask]

    # Remove specific increase columns from Total DataFrame if present
    columns_to_remove = ["Increase in Trade EBITDA", "Increase in Non-Trade EBITDA", "Increase in Total EBITDA"]
    columns_to_remove = [col for col in columns_to_remove if col in df_total.columns]
    df_total = df_total.drop(columns=columns_to_remove, errors='ignore')
    
    columns_to_remove = [col for col in columns_to_remove if col in df_no_total.columns]
    df_no_total = df_no_total.drop(columns=columns_to_remove, errors='ignore')
    
    return df_no_total, df_total

def streamlit_data_merger():
    # Set page configuration
    st.set_page_config(
        page_title="Data Merger App", 
        page_icon=":bar_chart:", 
        layout="wide"
    )

    # Custom CSS for styling
    st.markdown("""
    <style>
    .main-title {
        font-size: 2.5rem;
        color: #2C3E50;
        text-align: center;
        margin-bottom: 30px;
    }
    .upload-section {
        background-color: #F0F4F8;
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 20px;
    }
    .stButton>button {
        background-color: #3498DB;
        color: white;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background-color: #2980B9;
        transform: scale(1.05);
    }
    .hidden-info {
        color: #7F8C8D;
        font-size: 0.9rem;
        margin-top: 10px;
    }
    </style>
    """, unsafe_allow_html=True)

    # Title
    st.markdown('<h1 class="main-title">📊 Data Merger Application</h1>', unsafe_allow_html=True)

    # File upload section
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    
    # Initialize session state for files and sheets
    if 'files' not in st.session_state:
        st.session_state.files = [None, None, None]
    if 'selected_sheets' not in st.session_state:
        st.session_state.selected_sheets = [None, None, None]
    if 'visible_columns' not in st.session_state:
        st.session_state.visible_columns = [[], [], []]

    # File upload and sheet selection
    file_types = ["Oct-Sep", "Sep-Aug", "Aug-Jul"]
    
    for i in range(3):
        st.subheader(f"Upload {file_types[i]} File")
        uploaded_file = st.file_uploader(
            f"Choose Excel file for {file_types[i]}", 
            type=['xlsx', 'xls'], 
            key=f"file_uploader_{i}"
        )
        
        if uploaded_file is not None:
            st.session_state.files[i] = uploaded_file
            
            # Load workbook and get sheet names
            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            sheet_names = wb.sheetnames
            
            # Sheet selection
            selected_sheet = st.selectbox(
                f"Select sheet for {file_types[i]}", 
                sheet_names, 
                key=f"sheet_selector_{i}"
            )
            st.session_state.selected_sheets[i] = selected_sheet

            # Get visible columns
            ws = wb[selected_sheet]
            visible_cols = get_visible_columns(ws)
            st.session_state.visible_columns[i] = visible_cols
            
            # Display visible and hidden column information
            wb.close()
            
            # Column visibility information
            st.markdown(f"<div class='hidden-info'>Visible Columns: {len(visible_cols)}</div>", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # Merge button
    if st.button("Merge Data", key="merge_button"):
        # Validate file uploads and sheet selections
        if all(st.session_state.files) and all(st.session_state.selected_sheets):
            processed_dfs = []
            processed_total_dfs = []
            
            try:
                for i in range(3):
                    # Read only visible columns
                    df = read_visible_columns(
                        st.session_state.files[i], 
                        st.session_state.selected_sheets[i]
                    )
                    
                    # Preprocess and process
                    df = preprocess_data(df)
                    df_no_total, df_total = process_and_merge(df, file_types[i])
                    
                    processed_dfs.append(df_no_total)
                    processed_total_dfs.append(df_total)

                # Merge non-total DataFrames
                final_df = processed_dfs[0]
                for df in processed_dfs[1:]:
                    df = df[["Region Name", "Material Name"] + [col for col in df.columns if col not in final_df.columns]]
                    final_df = pd.merge(final_df, df, on=["Region Name", "Material Name"], how="left")
                
                # Merge total DataFrames
                final_total_df = processed_total_dfs[0]
                for df in processed_total_dfs[1:]:
                    df = df[["Region Name"] + [col for col in df.columns if col not in final_total_df.columns]]
                    final_total_df = pd.merge(final_total_df, df, on="Region Name", how="left")
                
                # Create Excel file in memory
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    final_df.to_excel(writer, sheet_name='Non-Total', index=False)
                    final_total_df.to_excel(writer, sheet_name='Total', index=False)
                output.seek(0)
                
                # Download button
                st.download_button(
                    label="Download Merged Excel File",
                    data=output,
                    file_name='final_merged_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Display DataFrames
                st.success("Data merged successfully!")
                
                # Columns information
                st.subheader("Merged Columns Information")
                col1, col2 = st.columns(2)
                
                with col1:
                    st.subheader("Non-Total DataFrame")
                    st.write("Columns:", final_df.columns.tolist())
                    st.dataframe(final_df.head())
                
                with col2:
                    st.subheader("Total DataFrame")
                    st.write("Columns:", final_total_df.columns.tolist())
                    st.dataframe(final_total_df.head())
                
            except Exception as e:
                st.error(f"An error occurred: {e}")
        else:
            st.warning("Please upload all three files and select sheets!")

# Run the Streamlit app
if __name__ == "__main__":
    streamlit_data_merger()
